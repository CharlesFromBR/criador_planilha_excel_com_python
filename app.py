import openpyxl

#criar uma planilha

book = openpyxl.Workbook()

#como visualizarpaginas existentes
print(book.sheetnames)

#como criar uma pagina
book.create_sheet('Aluno1')

#como selecionar uma pagina
frutas_page = book['Aluno1']
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Histoco','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])
frutas_page.append(['Historico','exemplo@gmail.com', 'nome completo'])

#salvara planilha
book.save('Planilha de Historico - Alunos.xlsx')